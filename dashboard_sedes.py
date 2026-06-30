import pandas as pd
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint

# ── helpers ────────────────────────────────────────────────────────────────────
def load_file(path, sede):
    sheets = pd.read_excel(path, sheet_name=None, header=None)
    df_raw = list(sheets.values())[0]
    header_row = None
    for i, row in df_raw.iterrows():
        if any(str(v).strip() == "Fecha" for v in row.values):
            header_row = i
            break
    df = df_raw.iloc[header_row+1:].copy()
    df.columns = range(len(df.columns))
    header = df_raw.iloc[header_row]
    col_map = {}
    for j, v in enumerate(header):
        sv = str(v).strip()
        if sv == "Fecha":          col_map["fecha"]      = j
        elif sv == "Cliente":      col_map["cliente"]    = j
        elif sv == "Registrado Por": col_map["vendedor"] = j
        elif sv == "Producto":     col_map["producto"]   = j
        elif sv == "Total CRC":    col_map["total_crc"]  = j
    result = pd.DataFrame()
    for key, col in col_map.items():
        result[key] = df[col].values
    result["sede"] = sede
    result = result.dropna(subset=["producto", "total_crc"])
    result["total_crc"] = pd.to_numeric(result["total_crc"], errors="coerce").fillna(0)
    result["producto"]  = result["producto"].astype(str).str.strip()
    result["vendedor"]  = result["vendedor"].astype(str).str.strip()
    result = result[(result["producto"] != "nan") & (result["total_crc"] > 0)]
    return result

MEMBERSHIP_KEYWORDS = [
    "plan", "clases", "nataci", "gym", "trimestre", "semestre", "anual",
    "sesion", "sesi", "renovaci", "activos", "verano", "black friday",
    "estimulaci", "aquafitness", "fit gym", "lealtad", "platiu", "platinum",
    "standar", "estandar", "exclusivo", "bloqueado", "multigym", "mensual",
    "estudiante", "adulto mayor", "corporativo", "aniversario", "inactivo",
    "promo", "clase baile", "clase grupal", "gym incluido", "nat amor",
    "ex nat", "ex 3 nat", "natacion adultos"
]

PERSONAL_TRAINER_KEYWORDS = ["personal training", "personal 2026"]

def classify(p):
    pl = p.lower()
    if any(k in pl for k in PERSONAL_TRAINER_KEYWORDS):
        return "Personal Trainer"
    if any(k in pl for k in MEMBERSHIP_KEYWORDS):
        return "Membresia"
    return "Minita"

# ── load & filter ──────────────────────────────────────────────────────────────
files = {
    "3 Rios":   r"C:\Users\Jose\Downloads\3rios.xlsx",
    "Natacion": r"C:\Users\Jose\Downloads\nata.xlsx",
    "Fita":     r"C:\Users\Jose\Downloads\fita.xlsx",
    "Saba":     r"C:\Users\Jose\Downloads\saba.xlsx",
}
dfs = [load_file(path, sede) for sede, path in files.items()]
df  = pd.concat(dfs, ignore_index=True)

# Exclude Yorleny
df = df[~df["vendedor"].str.contains("Yorleny", case=False, na=False)]
df["categoria"] = df["producto"].apply(classify)
# Personal Trainer stays in df (counts in total) but NOT in membresías

SEDES = ["3 Rios", "Natacion", "Fita", "Saba"]

# ── compute summaries ──────────────────────────────────────────────────────────
sede_summary = {}
for sede in SEDES:
    s = df[df["sede"] == sede]
    total  = s["total_crc"].sum()
    memb   = s[s["categoria"] == "Membresia"]["total_crc"].sum()
    minita = s[s["categoria"] == "Minita"]["total_crc"].sum()
    pt     = s[s["categoria"] == "Personal Trainer"]["total_crc"].sum()
    vend   = (s.groupby("vendedor")["total_crc"].sum()
               .sort_values(ascending=False).reset_index())
    vend.columns = ["Vendedor", "Total CRC"]
    sede_summary[sede] = dict(total=total, membresia=memb, minita=minita, personal_trainer=pt, vendedores=vend)

# ── Excel styling helpers ──────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MED_BLUE    = "2E5FA3"
LIGHT_BLUE  = "BDD7EE"
ACCENT_GOLD = "C9A84C"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
DARK_GRAY   = "404040"
GREEN       = "375623"
GREEN_LIGHT = "E2EFDA"
ORANGE_LIGHT= "FCE4D6"
ORANGE      = "C55A11"

SEDE_COLORS = {
    "3 Rios":   ("2E5FA3", "DEEAF1"),
    "Natacion": ("375623", "E2EFDA"),
    "Fita":     ("7030A0", "EAD1F7"),
    "Saba":     ("C55A11", "FCE4D6"),
}

thin = Side(style="thin", color="CCCCCC")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, cell, value, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, halign="center", valign="center"):
    c = ws[cell]
    c.value = value
    c.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=True)
    return c

def val(ws, cell, value, bg=WHITE, fg=DARK_GRAY, bold=False, size=10,
        halign="right", fmt=None, border=True):
    c = ws[cell]
    c.value = value
    c.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical="center")
    if border:
        c.border = thin_border
    if fmt:
        c.number_format = fmt
    return c

def merge_hdr(ws, r1, c1, r2, c2, text, bg=DARK_BLUE, fg=WHITE, size=12, bold=True):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(r1, c1)
    c.value = text
    c.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── SHEET 1: RESUMEN GENERAL ──────────────────────────────────────────────────
ws = wb.active
ws.title = "Resumen General"
ws.sheet_view.showGridLines = False

# title banner
ws.row_dimensions[1].height = 15
ws.row_dimensions[2].height = 38
ws.row_dimensions[3].height = 18
ws.merge_cells("A2:L2")
c = ws["A2"]
c.value = "REPORTE DE FACTURACIÓN POR SEDE – MAYO 2026"
c.font = Font(name="Calibri", bold=True, color=WHITE, size=18)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A3:L3")
c = ws["A3"]
c.value = "Excluye: Planes de Personal Training · Factura Yorleny (CRC 4,000,000)"
c.font = Font(name="Calibri", italic=True, color="888888", size=9)
c.fill = PatternFill("solid", fgColor="F0F4FA")
c.alignment = Alignment(horizontal="center", vertical="center")

# section header row 5
ROW_SH = 5
ws.row_dimensions[ROW_SH].height = 22
for col, text in enumerate([
    "SEDE", "TOTAL FACTURACIÓN", "MEMBRESÍAS", "% MEMBRESÍAS",
    "MINITA", "% MINITA", "PERSONAL TRAINER", "#1 VENDEDOR", "MONTO #1"
], start=2):
    hdr(ws, f"{get_column_letter(col)}{ROW_SH}", text, size=9, bg=MED_BLUE)

ws.row_dimensions[ROW_SH-1].height = 8  # spacer

# data rows
for r, sede in enumerate(SEDES, start=ROW_SH+1):
    ws.row_dimensions[r].height = 22
    s = sede_summary[sede]
    dark, light = SEDE_COLORS[sede]
    top_vend = s["vendedores"].iloc[0]["Vendedor"] if len(s["vendedores"]) > 0 else ""
    top_amt  = s["vendedores"].iloc[0]["Total CRC"] if len(s["vendedores"]) > 0 else 0
    pct_m    = s["membresia"] / s["total"] if s["total"] else 0
    pct_min  = s["minita"]    / s["total"] if s["total"] else 0

    bg = light if r % 2 == 0 else WHITE
    val(ws, f"B{r}", sede,                    bg=bg, fg=dark,      bold=True, halign="left")
    val(ws, f"C{r}", s["total"],              bg=bg, fg=DARK_GRAY,            fmt='#,##0')
    val(ws, f"D{r}", s["membresia"],          bg=bg, fg=DARK_GRAY,            fmt='#,##0')
    val(ws, f"E{r}", pct_m,                   bg=bg, fg=MED_BLUE, bold=True,  fmt='0.0%')
    val(ws, f"F{r}", s["minita"],             bg=bg, fg=DARK_GRAY,            fmt='#,##0')
    val(ws, f"G{r}", pct_min,                 bg=bg, fg=ORANGE,   bold=True,  fmt='0.0%')
    val(ws, f"H{r}", s["personal_trainer"],   bg=bg, fg=DARK_GRAY,            fmt='#,##0')
    val(ws, f"I{r}", top_vend,                bg=bg, fg=DARK_GRAY, halign="left")
    val(ws, f"J{r}", top_amt,                 bg=bg, fg=DARK_GRAY,            fmt='#,##0')

# totals row
r_tot = ROW_SH + len(SEDES) + 1
ws.row_dimensions[r_tot].height = 24
tot_total = sum(s["total"]     for s in sede_summary.values())
tot_memb  = sum(s["membresia"] for s in sede_summary.values())
tot_min   = sum(s["minita"]    for s in sede_summary.values())
tot_pt = sum(s["personal_trainer"] for s in sede_summary.values())
val(ws, f"B{r_tot}", "TOTAL GENERAL",      bg=DARK_BLUE, fg=WHITE, bold=True, halign="center", border=True)
val(ws, f"C{r_tot}", tot_total,            bg=DARK_BLUE, fg=WHITE, bold=True, fmt='#,##0')
val(ws, f"D{r_tot}", tot_memb,             bg=DARK_BLUE, fg=WHITE, bold=True, fmt='#,##0')
val(ws, f"E{r_tot}", tot_memb/tot_total,   bg=DARK_BLUE, fg=WHITE, bold=True, fmt='0.0%')
val(ws, f"F{r_tot}", tot_min,              bg=DARK_BLUE, fg=WHITE, bold=True, fmt='#,##0')
val(ws, f"G{r_tot}", tot_min/tot_total,    bg=DARK_BLUE, fg=WHITE, bold=True, fmt='0.0%')
val(ws, f"H{r_tot}", tot_pt,              bg=DARK_BLUE, fg=WHITE, bold=True, fmt='#,##0')
val(ws, f"I{r_tot}", "",                  bg=DARK_BLUE, fg=WHITE, bold=True)
val(ws, f"J{r_tot}", "",                  bg=DARK_BLUE, fg=WHITE, bold=True)

# column widths
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 16
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 16
ws.column_dimensions["I"].width = 30
ws.column_dimensions["J"].width = 18
ws.column_dimensions["K"].width = 2

# ── BAR CHART: total per sede ──────────────────────────────────────────────────
chart_row = r_tot + 3

# write invisible data range for chart
data_start = chart_row
for i, sede in enumerate(SEDES):
    ws.cell(data_start+i, 1, sede)
    ws.cell(data_start+i, 2, sede_summary[sede]["membresia"])
    ws.cell(data_start+i, 3, sede_summary[sede]["minita"])

bc = BarChart()
bc.type = "col"
bc.title = "Facturación por Sede"
bc.style = 10
bc.y_axis.title = "CRC"
bc.x_axis.title = "Sede"
bc.height = 12
bc.width  = 20
bc.grouping = "stacked"
bc.overlap  = 100

cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_start+3)
memb_ref = Reference(ws, min_col=2, min_row=data_start, max_row=data_start+3)
mini_ref = Reference(ws, min_col=3, min_row=data_start, max_row=data_start+3)
from openpyxl.chart import Series
s1 = Series(memb_ref, title="Membresías")
s2 = Series(mini_ref, title="Minita")
bc.series = [s1, s2]
bc.set_categories(cats)

ws.add_chart(bc, f"B{chart_row}")

# hide the raw data used for charts (make font white on white)
for i in range(4):
    for col in [1, 2, 3]:
        c = ws.cell(data_start+i, col)
        c.font = Font(color=WHITE)
        c.fill = PatternFill("solid", fgColor=WHITE)

# ── SHEET 2: VENDEDORES ───────────────────────────────────────────────────────
ws2 = wb.create_sheet("Vendedores por Sede")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 2

ws2.merge_cells("B1:M1")
c = ws2["B1"]
c.value = "RANKING DE VENDEDORES POR SEDE – MAYO 2026"
c.font = Font(name="Calibri", bold=True, color=WHITE, size=16)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 34

col_offsets = {"3 Rios": 2, "Natacion": 5, "Fita": 8, "Saba": 11}

for sede, start_col in col_offsets.items():
    dark, light = SEDE_COLORS[sede]
    vend_df = sede_summary[sede]["vendedores"]
    total_sede = sede_summary[sede]["total"]

    # Sede header spanning 2 cols
    ws2.merge_cells(start_row=3, start_column=start_col,
                    end_row=3,   end_column=start_col+1)
    c = ws2.cell(3, start_col)
    c.value = sede.upper()
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=dark)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[3].height = 22

    # sub-headers
    for ci, text in enumerate(["Vendedor", "CRC"], start=start_col):
        c = ws2.cell(4, ci)
        c.value = text
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=MED_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
    ws2.row_dimensions[4].height = 18

    for ri, (_, row) in enumerate(vend_df.iterrows()):
        r = 5 + ri
        bg = light if ri % 2 == 0 else WHITE
        ws2.row_dimensions[r].height = 18

        nombre = row["Vendedor"]
        monto  = row["Total CRC"]
        pct    = monto / total_sede if total_sede else 0

        c = ws2.cell(r, start_col)
        c.value = nombre
        c.font = Font(name="Calibri", size=9, color=DARK_GRAY,
                      bold=(ri == 0))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border

        c2 = ws2.cell(r, start_col+1)
        c2.value = monto
        c2.font = Font(name="Calibri", size=9, color=dark, bold=(ri == 0))
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c2.number_format = '#,##0'
        c2.border = thin_border

    # rank 1 gold highlight
    for ci in [start_col, start_col+1]:
        c = ws2.cell(5, ci)
        c.fill = PatternFill("solid", fgColor="FFF2CC")
        c.font = Font(name="Calibri", size=9, bold=True, color="7D6608")
        c.border = Border(
            left=Side(style="medium", color=ACCENT_GOLD),
            right=Side(style="medium", color=ACCENT_GOLD),
            top=Side(style="medium", color=ACCENT_GOLD),
            bottom=Side(style="medium", color=ACCENT_GOLD),
        )

# column widths ws2
for sede, start_col in col_offsets.items():
    ws2.column_dimensions[get_column_letter(start_col)].width   = 28
    ws2.column_dimensions[get_column_letter(start_col+1)].width = 15
    if start_col > 2:
        ws2.column_dimensions[get_column_letter(start_col-1)].width = 2  # gap

# ── SHEET 3: DETALLE POR SEDE ─────────────────────────────────────────────────
for sede in SEDES:
    dark, light = SEDE_COLORS[sede]
    ws3 = wb.create_sheet(sede)
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 2

    ws3.merge_cells("B1:H1")
    c = ws3["B1"]
    c.value = f"DETALLE DE VENTAS – {sede.upper()} – MAYO 2026"
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    c.fill = PatternFill("solid", fgColor=dark)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    # KPI cards row 3-5
    s = sede_summary[sede]
    kpis = [
        ("TOTAL FACTURACIÓN",  s["total"]),
        ("MEMBRESÍAS",         s["membresia"]),
        ("MINITA",             s["minita"]),
        ("PERSONAL TRAINER",   s["personal_trainer"]),
    ]
    for ki, (label, amount) in enumerate(kpis):
        col = 2 + ki*2
        for merge_r in [3, 4, 5]:
            ws3.merge_cells(start_row=merge_r, start_column=col, end_row=merge_r, end_column=col+1)

        c = ws3.cell(3, col)
        c.value = label
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=dark)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[3].height = 18

        c = ws3.cell(4, col)
        c.value = amount
        c.font = Font(name="Calibri", bold=True, color=dark, size=16)
        c.fill = PatternFill("solid", fgColor=light)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '#,##0'
        ws3.row_dimensions[4].height = 30

        pct = amount / s["total"] if s["total"] and label != "TOTAL FACTURACIÓN" else None
        c = ws3.cell(5, col)
        c.value = pct if pct is not None else ""
        c.font = Font(name="Calibri", color="888888", size=9)
        c.fill = PatternFill("solid", fgColor=light)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '0.0%'
        ws3.row_dimensions[5].height = 16

    # table header row 7
    cols_detail = ["Fecha", "Cliente", "Vendedor", "Producto", "Categoría", "Total CRC"]
    ws3.row_dimensions[6].height = 6
    ws3.row_dimensions[7].height = 20
    for ci, col_name in enumerate(cols_detail, start=2):
        c = ws3.cell(7, ci)
        c.value = col_name
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=dark)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    # data
    sede_df = df[df["sede"] == sede].copy()
    for ri, (_, row) in enumerate(sede_df.iterrows()):
        r = 8 + ri
        ws3.row_dimensions[r].height = 16
        bg = light if ri % 2 == 0 else WHITE

        fecha_val = row.get("fecha", "")
        if hasattr(fecha_val, 'strftime'):
            fecha_str = fecha_val.strftime("%d/%m/%Y")
        else:
            fecha_str = str(fecha_val)[:10] if fecha_val else ""

        row_data = [fecha_str, str(row.get("cliente",""))[:40],
                    row["vendedor"], row["producto"],
                    row["categoria"], row["total_crc"]]
        for ci, val_d in enumerate(row_data, start=2):
            c = ws3.cell(r, ci)
            c.value = val_d
            c.font = Font(name="Calibri", size=8, color=DARK_GRAY)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = thin_border
            if ci == 7:  # Total CRC
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    ws3.column_dimensions["A"].width = 2
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 28
    ws3.column_dimensions["D"].width = 28
    ws3.column_dimensions["E"].width = 38
    ws3.column_dimensions["F"].width = 12
    ws3.column_dimensions["G"].width = 14

    ws3.freeze_panes = "B8"

# ── save ──────────────────────────────────────────────────────────────────────
out = r"C:\Users\Jose\Downloads\Dashboard_Sedes_Mayo2026.xlsx"
wb.save(out)
print(f"Guardado: {out}")
